#type: ignore
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = openpyxl.Workbook()
worksheet: Worksheet = workbook.active

#Criando calendários
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

students = [
    ['João', 14, 5.5],
    ['Maria', 12, 7],
    ['Rebeca', 20, 10],
    ['Antônio', 44, 8],
]

workbook.save(WORKBOOK_PATH)    
